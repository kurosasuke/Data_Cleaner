import numpy as np
import openpyxl
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit

# Get Excel data
worksheet = openpyxl.load_workbook(filename=r"C:\Users\darii\OneDrive\Desktop\FC_data.xlsx")

sheet = worksheet.active
col_T = "C"
col_V = "D"
col_C = "E"

Time = []
Volt = []
Curr = []

# Import Seconds from Excel
for i in range(len(sheet[col_T])):
    if isinstance(sheet.cell(row=i + 1, column=3).value, int):
        val_t = sheet.cell(row=i + 1, column=3).value
        Time.append(val_t)

# Import Voltage from Excel
for i in range(len(sheet[col_V])):
    if isinstance(sheet.cell(row=i + 1, column=4).value, float) or \
            isinstance(sheet.cell(row=i + 1, column=4).value, int):
        val_v = sheet.cell(row=i + 1, column=4).value
        Volt.append(val_v)

# Import Current from Excel
for i in range(len(sheet[col_C])):
    if isinstance(sheet.cell(row=i + 1, column=5).value, float) or \
            isinstance(sheet.cell(row=i + 1, column=5).value, int):
        val_c = sheet.cell(row=i + 1, column=5).value
        Curr.append(val_c)

# Time start from zero
Time = Time[250:]
Volt = Volt[250:]
Curr = Curr[250:]

# Clean data from zeros
Volt_n = []
Curr_n = []

for i in range(len(Volt)):
    if Volt[i] != 0:
        Volt_n.append(Volt[i])
        Curr_n.append(Curr[i])

# Ramping variables
new_Time = []
fake_time = []

# Distance between two ramping sets
for i in range(len(Time)):
    if Time[i - 1] != 0 and Time[i] == 0 and Time[i + 1] == 0:
        new_Time.append(fake_time)
        fake_time = []
    else:
        fake_time.append(Time[i])

# Delete I<40A
new_Volt = []
new_Curr = []

for i in range(len(Volt_n)):
    if Volt_n[i] > 40:
        new_Volt.append(Volt_n[i])
        new_Curr.append(Curr_n[i])

# Delete scatter datas
Volt_r = []
Curr_r = []

for i in range(len(new_Volt)-1):
    if new_Volt[i+1] - new_Volt[i] < 0.2:
        Volt_r.append(new_Volt[i])
        Curr_r.append(new_Curr[i])

# Plot results
plt.scatter(Volt_r, Curr_r)

# Trend-line
z = np.polyfit(Volt_r, Curr_r, 5)
p = np.poly1d(z)

# Add trend-line to scatter plot
xmodel = np.arange(40, 70, 0.1)
ymodel = np.polyval(z, xmodel)
plt.plot(xmodel, ymodel, color="purple", linewidth=3, linestyle="--")
plt.title("Polarization function (V-I)")
plt.xlabel("Voltage [V]")
plt.ylabel("Current [A]")

# Show plot on Pycharm (maybe you doesn't need it)
plt.show()
