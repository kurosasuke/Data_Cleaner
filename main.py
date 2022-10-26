import matplotlib.pyplot as plt
import numpy
import openpyxl

# Get Excel data
worksheet = openpyxl.load_workbook(filename=r"C:\Users\darii\OneDrive\Desktop\FC_data.xlsx")

workbook = openpyxl.Workbook()
newworksheet = workbook.active

sheet = worksheet.active
col = "D"

# initialization
i = 1
xxx = numpy.zeros(len(sheet[col]) + 1)
yyy = numpy.zeros(len(sheet[col]) + 1)

# Get Voltage column position
for k in range(7):
    if sheet.cell(row=1, column=k + 1).value == "Voltage":
        p_volt = k+1

# Delete short circuit values
for x in range(len(sheet[col])):
    s = x + 1
    if sheet.cell(row=s, column=p_volt).value == 0 and sheet.cell(row=s - 1, column=p_volt).value != 0 and \
            sheet.cell(row=s + 1, column=p_volt).value != 0:
        print("")
    else:
        newworksheet.cell(row=i, column=1).value = sheet.cell(row=s, column=p_volt).value
        newworksheet.cell(row=i, column=2).value = sheet.cell(row=s, column=p_volt+1).value
        if type(newworksheet.cell(row=i, column=1).value) != str:
            xxx[s] = sheet.cell(row=s, column=4).value
            yyy[s] = sheet.cell(row=s, column=5).value
        i += 1

# Save right values
workbook.save("FC_Data_Cleaned.xlsx")

# Plot results
plt.scatter(xxx, yyy)
plt.show()
